"""Inspect Montana.xlsx - find headers, sample raw township values."""
from pathlib import Path
from collections import Counter
import openpyxl

wb = openpyxl.load_workbook(r"G:\My Drive\Agriculture\Monette\Montana.xlsx",
                             data_only=True, read_only=True)
print(f"sheets: {wb.sheetnames}")
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"\n=== sheet: {sn} ===  (rows: {ws.max_row}, cols: {ws.max_column})")
    # print first 4 rows
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 8:
            break
        print(f"  row {i}: {row[:15]}")

    # find the actual header row (containing 'Township')
    hdr_idx = None
    hdr = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if any(isinstance(c, str) and c.strip().lower() == "township" for c in (row or [])):
            hdr_idx = i
            hdr = row
            break
    if hdr is None:
        print("  NO HEADER with 'Township' found")
        continue
    print(f"  header row: {hdr_idx}")
    print(f"  headers: {hdr}")

    c_twp = next(i for i, h in enumerate(hdr) if isinstance(h, str) and h.strip().lower() == "township")
    c_rng = next(i for i, h in enumerate(hdr) if isinstance(h, str) and h.strip().lower() == "range")
    c_sec = next(i for i, h in enumerate(hdr) if isinstance(h, str) and h.strip().lower() == "section")
    c_county = next((i for i, h in enumerate(hdr) if isinstance(h, str) and "county" in h.strip().lower() and "cd" not in h.strip().lower()), None)
    c_propname = next((i for i, h in enumerate(hdr) if isinstance(h, str) and "propertyname" in h.strip().lower().replace(" ", "")), None)
    c_tax = next((i for i, h in enumerate(hdr) if isinstance(h, str) and "taxyear" in h.strip().lower().replace(" ", "")), None)

    print(f"  col indices: twp={c_twp}, rng={c_rng}, sec={c_sec}, county={c_county}, propname={c_propname}, tax={c_tax}")

    twp_values = Counter()
    rng_values = Counter()
    distinct_parcels = set()
    distinct_parcels_with_year = set()
    county_counts = Counter()
    propname_counts = Counter()
    tax_counts = Counter()

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= hdr_idx:
            continue
        if row is None or len(row) <= max(c_twp, c_rng, c_sec):
            continue
        tw = row[c_twp]
        rn = row[c_rng]
        sc = row[c_sec]
        if tw is None and rn is None and sc is None:
            continue
        twp_values[str(tw)] += 1
        rng_values[str(rn)] += 1
        if c_county is not None and c_county < len(row):
            county_counts[str(row[c_county])] += 1
        if c_propname is not None and c_propname < len(row):
            propname_counts[str(row[c_propname])] += 1
        if c_tax is not None and c_tax < len(row):
            tax_counts[str(row[c_tax])] += 1
        distinct_parcels.add((str(tw), str(rn), str(sc)))
        distinct_parcels_with_year.add((str(tw), str(rn), str(sc), str(row[c_tax]) if c_tax is not None and c_tax < len(row) else ""))

    print(f"\n  distinct Twp: {dict(twp_values.most_common())}")
    print(f"\n  distinct Rng: {dict(rng_values.most_common())}")
    print(f"\n  distinct (twp,rng,sec): {len(distinct_parcels)}")
    print(f"  distinct (twp,rng,sec,year): {len(distinct_parcels_with_year)}")
    print(f"\n  county distribution: {dict(county_counts.most_common())}")
    if c_propname is not None:
        print(f"\n  propname distribution: {dict(propname_counts.most_common())}")
    if c_tax is not None:
        print(f"\n  taxyear distribution: {dict(tax_counts.most_common())}")

wb.close()

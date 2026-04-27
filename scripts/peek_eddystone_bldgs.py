"""Show the full buildings sheet + a C-of-T clustering on third-party rows.
Goal: figure out how many distinct third-party entities are likely at Eddystone."""
from __future__ import annotations
from pathlib import Path
import openpyxl

wb = openpyxl.load_workbook(r"G:\My Drive\Agriculture\Monette\Eddystone.xlsx",
                            data_only=True, read_only=True)

print("=== FULL BUILDINGS SHEET ===")
ws2 = wb["13-Eddystone_Bldgs"]
for r_idx, row in enumerate(ws2.iter_rows(values_only=True), start=1):
    vals = [c for c in row if c is not None]
    if vals:
        print(f"row {r_idx:3d}: {list(row)[:10]}")

print("\n=== C OF T CLUSTERING ON 3rd-PARTY ROWS ===")
ws = wb["13-Eddystone"]
rows = list(ws.iter_rows(values_only=True))[2:]
third = [r for r in rows if r[3] == "3rd Party Vendor"]
print(f"Total 3rd-party rows: {len(third)}")

cots = sorted((r[2] for r in third if isinstance(r[2], int)))
print(f"\nAll C of T numbers (sorted, {len(cots)} total):")
# cluster by rough range (100k windows)
import collections
buckets = collections.Counter()
for cot in cots:
    buckets[(cot // 100000) * 100000] += 1
for base, cnt in sorted(buckets.items()):
    print(f"  {base:,}–{base+99_999:,}: {cnt} titles")

print(f"\nAll third-party C of T values (raw):")
for cot in cots:
    print(f"  {cot}")

# Group 3rd-party parcels by township — does any township contain ONLY
# third-party parcels (ie a contiguous bloc) vs mixed with Monette?
print("\n=== TWP ROLLUP: MONETTE vs 3rd-PARTY ACRES ===")
from collections import defaultdict
by_twp = defaultdict(lambda: {"mon":0.0, "tp":0.0, "mon_ct":0, "tp_ct":0})
for r in rows:
    community = r[11] or ""
    size = r[7] if isinstance(r[7], (int, float)) else 0.0
    if r[3] == "Monette Farms":
        by_twp[community]["mon"] += size
        by_twp[community]["mon_ct"] += 1
    elif r[3] == "3rd Party Vendor":
        by_twp[community]["tp"] += size
        by_twp[community]["tp_ct"] += 1

for twp, stats in sorted(by_twp.items()):
    print(f"  {twp:30s}  Monette={stats['mon']:>8,.0f}ac ({stats['mon_ct']:3d})  3rd-Party={stats['tp']:>8,.0f}ac ({stats['tp_ct']:3d})")

# Civic addresses on third-party rows — may give building locations
print("\n=== CIVIC ADDRESSES ON 3rd-PARTY ROWS (non-blank) ===")
for r in third:
    if r[4]:  # civic address
        print(f"  COT={r[2]}  LLD={r[10]:20s}  ADDR={r[4]!r}")

wb.close()

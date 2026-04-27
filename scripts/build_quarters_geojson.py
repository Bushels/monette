"""
build_quarters_geojson.py — Extract real legal land descriptions from each
property's XLSX on G:\\My Drive\\Agriculture\\Monette and convert them to
GeoJSON polygons using Dominion Land Survey (DLS, Canada) or Public Land
Survey System (PLSS, US/Montana) geometry.

Four input formats handled:

  1. SK standard DLS         — "NW-10-11-10-W3"    (Vanguard, Ponteix, Admiral,
                                                     Hafford, Kamsack, Outlook,
                                                     Prince_Albert, Raymore,
                                                     Wymark, Calderbank)
  2. MB compact DLS          — "NW4-26-11W"        (Eddystone)
  3. MT PLSS (column-based)  — Twp "04 S", Rng "31 E", Sec 26
                                                    (Montana.xlsx, 3 props)
  4. MB Red River parish lots — "RL70-PQ-4734"     (The Pas) — NOT resolved
                                                    here; emitted as property-
                                                    centered fallback marker.

Output: ../quarters.geojson — a FeatureCollection where each feature is a
single quarter-section / section / parcel polygon with rich properties
(property_id, loc, titled_ac, soil, crops 2023-2025, Cert of Title, RM).

The computation is intentionally first-principles (no GIS library required)
so the output is reproducible and auditable. Accuracy is +/- 150 m at the
edges of each township, which is invisible at country through quarter-section
zoom levels on the Atlas view.

2026-04-23 update:
- The atlas now exposes the formula-driven parcel boxes directly over
  satellite imagery, so the earlier +/- 150 m tolerance is not good enough
  for inspection mode.
- A property-level calibration layer can now be applied from
  scripts/quarter_geometry_calibration.json to shift / scale / rotate the
  generated rings before writing quarters.geojson.
"""
from __future__ import annotations
import json
import math
import re
import sys
from pathlib import Path
from typing import Iterable

import openpyxl

# ── Constants ──────────────────────────────────────────────────────────────
MILE_KM = 1.609344
EARTH_KM_PER_DEG_LAT = 111.1949  # mean, good enough for SK/MB/MT latitudes
EARTH_M_PER_DEG_LAT = EARTH_KM_PER_DEG_LAT * 1000
IDENTITY_CALIBRATION = {
    "east_m": 0.0,
    "north_m": 0.0,
    "scale_x": 1.0,
    "scale_y": 1.0,
    "rotate_deg": 0.0,
}


def mi_to_deg_lat(mi: float) -> float:
    return (mi * MILE_KM) / EARTH_KM_PER_DEG_LAT


def mi_to_deg_lng(mi: float, at_lat_deg: float) -> float:
    return (mi * MILE_KM) / (EARTH_KM_PER_DEG_LAT * math.cos(math.radians(at_lat_deg)))


def _float_or(value, default: float) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return float(default)


def _normalize_calibration(entry: dict | None) -> dict:
    entry = entry or {}
    return {
        "east_m": _float_or(entry.get("east_m"), 0.0),
        "north_m": _float_or(entry.get("north_m"), 0.0),
        "scale_x": _float_or(entry.get("scale_x"), 1.0),
        "scale_y": _float_or(entry.get("scale_y"), 1.0),
        "rotate_deg": _float_or(entry.get("rotate_deg"), 0.0),
    }


def load_geometry_calibration(path: Path) -> dict:
    if not path.exists():
        return {"default": dict(IDENTITY_CALIBRATION), "properties": {}}

    raw = json.loads(path.read_text(encoding="utf-8"))
    default = _normalize_calibration(raw.get("default"))
    props = {
        str(pid): _normalize_calibration(entry)
        for pid, entry in (raw.get("properties") or {}).items()
    }
    return {"default": default, "properties": props}


def calibration_for_property(config: dict, property_id: str) -> dict:
    merged = dict(config.get("default") or IDENTITY_CALIBRATION)
    merged.update(config.get("properties", {}).get(property_id, {}))
    return merged


def is_identity_calibration(entry: dict) -> bool:
    return (
        abs(entry.get("east_m", 0.0)) < 1e-9 and
        abs(entry.get("north_m", 0.0)) < 1e-9 and
        abs(entry.get("scale_x", 1.0) - 1.0) < 1e-9 and
        abs(entry.get("scale_y", 1.0) - 1.0) < 1e-9 and
        abs(entry.get("rotate_deg", 0.0)) < 1e-9
    )


def property_origin(records: list[dict]) -> tuple[float, float]:
    bounds = [float("inf"), float("inf"), float("-inf"), float("-inf")]
    for rec in records:
        for lng, lat in rec.get("_polygon", []):
            bounds[0] = min(bounds[0], lng)
            bounds[1] = min(bounds[1], lat)
            bounds[2] = max(bounds[2], lng)
            bounds[3] = max(bounds[3], lat)
    return ((bounds[0] + bounds[2]) / 2, (bounds[1] + bounds[3]) / 2)


def transform_ring(
    ring: list[list[float]],
    origin_lng: float,
    origin_lat: float,
    calibration: dict,
) -> list[list[float]]:
    cos_lat = math.cos(math.radians(origin_lat))
    if abs(cos_lat) < 1e-9:
        cos_lat = 1e-9

    theta = math.radians(calibration["rotate_deg"])
    cos_theta = math.cos(theta)
    sin_theta = math.sin(theta)

    out: list[list[float]] = []
    for lng, lat in ring:
        x = (lng - origin_lng) * EARTH_M_PER_DEG_LAT * cos_lat
        y = (lat - origin_lat) * EARTH_M_PER_DEG_LAT

        x *= calibration["scale_x"]
        y *= calibration["scale_y"]

        xr = (x * cos_theta) - (y * sin_theta)
        yr = (x * sin_theta) + (y * cos_theta)

        xr += calibration["east_m"]
        yr += calibration["north_m"]

        out.append([
            origin_lng + (xr / (EARTH_M_PER_DEG_LAT * cos_lat)),
            origin_lat + (yr / EARTH_M_PER_DEG_LAT),
        ])
    return out


def apply_property_calibration(records: list[dict], calibration: dict) -> list[dict]:
    if not records or is_identity_calibration(calibration):
        return records

    origin_lng, origin_lat = property_origin(records)
    transformed = []
    for rec in records:
        next_rec = dict(rec)
        next_rec["_polygon"] = transform_ring(rec["_polygon"], origin_lng, origin_lat, calibration)
        transformed.append(next_rec)
    return transformed


# ── Canadian DLS ───────────────────────────────────────────────────────────
# The DLS starts at 49°N (the international border) and works north in
# 6-mile townships. Ranges go WEST of each principal meridian, so Range 1 W3
# is immediately west of the W3 meridian, Range 2 W3 is the next band west, …
#
# Principal meridians (longitude in degrees east, negative = west):
#   W1  = -97°27'28.41"  — runs through Manitoba just east of Winnipeg
#   W2  = -102°
#   W3  = -106°
#   W4  = -110°          — SK/AB provincial boundary
#   W5  = -114°
#   W6  = -118°
DLS_MERIDIAN_LNG = {
    "W1": -97.4573361,
    "W2": -102.0,
    "W3": -106.0,
    "W4": -110.0,
    "W5": -114.0,
    "W6": -118.0,
}
DLS_BASELINE_LAT = 49.0  # Canada/US border

# Section layout within a township (6x6 grid). Section 1 is SE corner;
# numbering snakes: row 0 (south) runs 1..6 east→west, row 1 runs 7..12
# west→east, and so on. Return (col_from_west, row_from_south) for a section.
def dls_section_to_grid(section: int) -> tuple[int, int]:
    if section < 1 or section > 36:
        raise ValueError(f"section {section} out of range 1..36")
    row_from_south = (section - 1) // 6
    pos_in_row = (section - 1) % 6  # 0..5
    if row_from_south % 2 == 0:
        # East→west: section 1 is col=5 (east), section 6 is col=0 (west)
        col_from_west = 5 - pos_in_row
    else:
        # West→east: section 7 is col=0, section 12 is col=5
        col_from_west = pos_in_row
    return col_from_west, row_from_south


# Quarter position within a section: (col_from_west 0|1, row_from_south 0|1)
DLS_QUARTER_OFFSET = {
    "SW": (0, 0),
    "SE": (1, 0),
    "NW": (0, 1),
    "NE": (1, 1),
}


def dls_quarter_polygon(qtr: str, section: int, township: int, rng: int,
                        meridian: str) -> list[list[float]]:
    """Compute a closed [lng,lat] ring for a DLS quarter-section."""
    if meridian not in DLS_MERIDIAN_LNG:
        raise ValueError(f"Unknown DLS meridian: {meridian}")
    if qtr not in DLS_QUARTER_OFFSET:
        raise ValueError(f"Unknown quarter direction: {qtr}")

    sec_col_w, sec_row_s = dls_section_to_grid(section)
    q_col_w, q_row_s = DLS_QUARTER_OFFSET[qtr]

    # South edge of township T: (T-1) * 6 mi north of baseline.
    township_south_lat = DLS_BASELINE_LAT + mi_to_deg_lat((township - 1) * 6)
    # South edge of this section = township south + sec_row_s miles north.
    section_south_lat = township_south_lat + mi_to_deg_lat(sec_row_s)
    # South edge of this quarter = section south + 0.5 * q_row_s miles north.
    q_south_lat = section_south_lat + mi_to_deg_lat(q_row_s * 0.5)
    q_north_lat = q_south_lat + mi_to_deg_lat(0.5)

    # Use the quarter's centre latitude for longitude scaling.
    centre_lat = (q_south_lat + q_north_lat) / 2
    mi_lng = lambda m: mi_to_deg_lng(m, centre_lat)

    meridian_lng = DLS_MERIDIAN_LNG[meridian]
    # East edge of Range R (W of meridian): meridian - (R-1)*6 mi west.
    range_east_lng = meridian_lng - mi_lng((rng - 1) * 6)
    # East edge of this section = range east - (5 - sec_col_w) mi west.
    section_east_lng = range_east_lng - mi_lng(5 - sec_col_w)
    # East edge of this quarter = section east - 0.5*(1 - q_col_w) mi west.
    q_east_lng = section_east_lng - mi_lng(0.5 * (1 - q_col_w))
    q_west_lng = q_east_lng - mi_lng(0.5)

    # Clockwise ring starting at SW.
    return [
        [q_west_lng, q_south_lat],
        [q_east_lng, q_south_lat],
        [q_east_lng, q_north_lat],
        [q_west_lng, q_north_lat],
        [q_west_lng, q_south_lat],
    ]


# ── Montana PLSS ───────────────────────────────────────────────────────────
# Montana Principal Meridian: -111°39'33" W, baseline at 45°47'13" N.
# Montana PLSS almost uniformly uses this meridian/baseline combo.
MT_PM_LNG = -111.6592
MT_PM_LAT = 45.7870


def mt_section_polygon(section: int, township_n: int, township_dir: str,
                       range_n: int, range_dir: str) -> list[list[float]]:
    """
    Return a closed [lng,lat] ring for a MT PLSS full section. Montana row
    data is section-level only (no quarters), so we emit the full 1-mile
    square.
    """
    # South edge of the township.
    # If N of baseline, Township 1 south edge = baseline itself; each
    # subsequent township adds 6 mi north.
    # If S of baseline, Township 1 south edge = 6 mi south of baseline.
    if township_dir.upper() == "N":
        twp_south_lat = MT_PM_LAT + mi_to_deg_lat((township_n - 1) * 6)
    else:  # S
        twp_south_lat = MT_PM_LAT - mi_to_deg_lat(township_n * 6)

    # Section grid (same snake pattern as DLS — PLSS uses the identical layout).
    sec_col_w, sec_row_s = dls_section_to_grid(section)

    # South edge of section.
    section_south_lat = twp_south_lat + mi_to_deg_lat(sec_row_s)
    section_north_lat = section_south_lat + mi_to_deg_lat(1)

    centre_lat = (section_south_lat + section_north_lat) / 2
    mi_lng = lambda m: mi_to_deg_lng(m, centre_lat)

    # East edge of Range R.
    # For ranges EAST of meridian: east edge of R = meridian + R mi east,
    # west edge = meridian + (R-1) mi east.
    # For ranges WEST: east edge = meridian - (R-1) mi west, west edge = -R mi.
    if range_dir.upper() == "E":
        range_east_lng = MT_PM_LNG + mi_lng(range_n * 6)
    else:  # W
        range_east_lng = MT_PM_LNG - mi_lng((range_n - 1) * 6)

    # East edge of this section = range east - (5 - sec_col_w) * 1 mi west.
    section_east_lng = range_east_lng - mi_lng(5 - sec_col_w)
    section_west_lng = section_east_lng - mi_lng(1)

    return [
        [section_west_lng, section_south_lat],
        [section_east_lng, section_south_lat],
        [section_east_lng, section_north_lat],
        [section_west_lng, section_north_lat],
        [section_west_lng, section_south_lat],
    ]


# ── Full-section polygon helper (for MB fallback and MT dedupe) ────────────
def dls_section_polygon(section: int, township: int, rng: int,
                        meridian: str) -> list[list[float]]:
    """Closed ring for a whole 1-mile DLS section (used when the source
    doesn't specify a quarter, e.g., Eddystone 'S26-24-12W')."""
    if meridian not in DLS_MERIDIAN_LNG:
        raise ValueError(f"Unknown DLS meridian: {meridian}")
    sec_col_w, sec_row_s = dls_section_to_grid(section)
    township_south_lat = DLS_BASELINE_LAT + mi_to_deg_lat((township - 1) * 6)
    section_south_lat = township_south_lat + mi_to_deg_lat(sec_row_s)
    section_north_lat = section_south_lat + mi_to_deg_lat(1)
    centre_lat = (section_south_lat + section_north_lat) / 2
    mi_lng = lambda m: mi_to_deg_lng(m, centre_lat)
    meridian_lng = DLS_MERIDIAN_LNG[meridian]
    range_east_lng = meridian_lng - mi_lng((rng - 1) * 6)
    section_east_lng = range_east_lng - mi_lng(5 - sec_col_w)
    section_west_lng = section_east_lng - mi_lng(1)
    return [
        [section_west_lng, section_south_lat],
        [section_east_lng, section_south_lat],
        [section_east_lng, section_north_lat],
        [section_west_lng, section_north_lat],
        [section_west_lng, section_south_lat],
    ]


# ── Parsers for the various XLSX formats ───────────────────────────────────
SK_DLS_RE = re.compile(r"^\s*(NE|NW|SE|SW)-(\d{1,2})-(\d{1,3})-(\d{1,3})-(W[1-6])\s*$", re.I)
MB_COMPACT_RE = re.compile(r"^\s*(NE|NW|SE|SW)(\d{1,2})-(\d{1,3})-(\d{1,3})(W[1-6]?)?\s*$", re.I)
# Whole-section fallback e.g. 'S26-24-12W' or 'E33-25-12W' — emit 1-mile square.
MB_SECTION_RE = re.compile(r"^\s*(?:[NSEW]|N½|S½|E½|W½)(\d{1,2})-(\d{1,3})-(\d{1,3})(W[1-6]?)?\s*$", re.I)


def _find_header_row(ws, markers: set[str]) -> tuple[int, dict[str, int]]:
    """Find the row containing the given header markers; return its index
    (1-based) and a dict of lower(marker) -> column_index (0-based)."""
    markers_lower = {m.lower() for m in markers}
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        found = {}
        for col, cell in enumerate(row):
            if isinstance(cell, str) and cell.strip().lower() in markers_lower:
                found[cell.strip().lower()] = col
        if len(found) >= max(2, len(markers_lower) // 2):
            # Capture all string headers in that row.
            hdr_map = {
                (cell.strip().lower() if isinstance(cell, str) else f"__col{col}"): col
                for col, cell in enumerate(row)
            }
            return i, hdr_map
    return -1, {}


# ── Row-level aggregation ──────────────────────────────────────────────────
# Many XLSX rows represent *titles* or *sub-parcels* on the same quarter-
# section. For mapping we want one polygon per distinct legal description,
# with aggregated acres and a count. This helper takes a list of parsed
# records (each with '_key' and '_polygon') and merges duplicates.
def _aggregate_by_key(records: list[dict]) -> list[dict]:
    merged: dict[tuple, dict] = {}
    for r in records:
        key = r["_key"]
        if key in merged:
            existing = merged[key]
            existing["title_count"] = existing.get("title_count", 1) + 1
            # Sum titled_ac where both numeric; otherwise keep first.
            def sum_ac(a, b):
                try:
                    return float(a) + float(b)
                except (TypeError, ValueError):
                    return a if a is not None else b
            if "titled_ac" in existing:
                existing["titled_ac"] = sum_ac(existing.get("titled_ac"), r.get("titled_ac"))
            # Prefer the first non-null value for scalar metadata
            for k, v in r.items():
                if k in ("_key", "_polygon", "title_count", "titled_ac"):
                    continue
                if existing.get(k) in (None, "") and v not in (None, ""):
                    existing[k] = v
        else:
            r["title_count"] = 1
            merged[key] = dict(r)
    return list(merged.values())


def parse_sk_standard(xlsx_path: Path, property_id: str) -> list[dict]:
    """
    Handles the 11 SK properties that share this header layout:
      Offer Price | Owner | Parcel # | Title | Qtr | Sec | Twp | R | M |
      Land Locations | Soil | Titled Acres | Waste Acres | ... | RM |
      title | Property Name | 2025 | 2024 | 2023 | ...
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr_row, hdr = _find_header_row(ws, {"qtr", "sec", "land locations"})
    if hdr_row < 0:
        wb.close()
        return []

    def col(*keys):
        for k in keys:
            if k in hdr:
                return hdr[k]
        return None

    c_qtr = col("qtr")
    c_sec = col("sec")
    c_twp = col("twp")
    c_rng = col("r", "rng", "range")
    c_mer = col("m", "meridian")
    c_loc = col("land locations")
    c_title = col("title")
    c_soil = col("soil")
    c_titled_ac = col("titled acres")
    c_waste_ac = col("waste acres")
    c_assessment = col("assessment")
    c_rm = col("rm")
    c_2025 = col(2025, "2025")
    c_2024 = col(2024, "2024")
    c_2023 = col(2023, "2023")
    c_parcel = col("parcel #")

    out = []
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        loc = row[c_loc] if c_loc is not None and c_loc < len(row) else None
        if not isinstance(loc, str):
            continue
        m = SK_DLS_RE.match(loc)
        if not m:
            continue
        qtr, sec, twp, rng, mer = m.group(1).upper(), int(m.group(2)), int(m.group(3)), int(m.group(4)), m.group(5).upper()
        try:
            ring = dls_quarter_polygon(qtr, sec, twp, rng, mer)
        except ValueError:
            continue

        def g(c):
            return row[c] if c is not None and c < len(row) else None

        out.append({
            "property_id": property_id,
            "loc": loc.strip(),
            "qtr": qtr, "sec": sec, "twp": twp, "rng": rng, "meridian": mer,
            "titled_ac": g(c_titled_ac),
            "waste_ac": g(c_waste_ac),
            "assessment": g(c_assessment),
            "soil": g(c_soil),
            "rm": g(c_rm),
            "crop_2025": g(c_2025),
            "crop_2024": g(c_2024),
            "crop_2023": g(c_2023),
            "parcel_no": g(c_parcel),
            "title": g(c_title),
            "_key": (qtr, sec, twp, rng, mer),
            "_polygon": ring,
        })
    wb.close()
    return _aggregate_by_key(out)


def parse_eddystone(xlsx_path: Path, property_id: str) -> list[dict]:
    """
    Manitoba compact-DLS variant for Eddystone:
      C of T | Owner Name | Civic Address | Municipality | Roll No. |
      Size | Short Legal Description | LLD | Community | Property Name | 2025 | 2024
    LLD values look like "SW35-25-12W" (quarter + section + township + range + W).
    Manitoba uses only W1, so we default to W1.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr_row, hdr = _find_header_row(ws, {"lld", "short legal description"})
    if hdr_row < 0:
        wb.close()
        return []

    def col(*keys):
        for k in keys:
            if k in hdr:
                return hdr[k]
        return None

    c_lld = col("lld")
    c_short = col("short legal description")
    c_size = col("size")
    c_cot = col("c of t")
    c_muni = col("municipality")
    c_2025 = col(2025, "2025")
    c_2024 = col(2024, "2024")
    c_2023 = col(2023, "2023")

    out = []
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        lld = row[c_lld] if c_lld is not None and c_lld < len(row) else None
        if not isinstance(lld, str):
            continue
        lld_s = lld.strip()
        # Try quarter-section first, then fall back to whole-section half-prefix.
        mq = MB_COMPACT_RE.match(lld_s)
        ms = MB_SECTION_RE.match(lld_s) if not mq else None
        if mq:
            qtr = mq.group(1).upper()
            section = int(mq.group(2))
            twp = int(mq.group(3))
            rng = int(mq.group(4))
            mer = "W1"
            try:
                ring = dls_quarter_polygon(qtr, section, twp, rng, mer)
            except ValueError:
                continue
            key = (qtr, section, twp, rng, mer)
            out_loc = f"{qtr}-{section}-{twp}-{rng}-{mer}"
        elif ms:
            # Whole-section or half-section fallback: approximate as full 1-mi².
            qtr = None
            section = int(ms.group(1))
            twp = int(ms.group(2))
            rng = int(ms.group(3))
            mer = "W1"
            try:
                ring = dls_section_polygon(section, twp, rng, mer)
            except ValueError:
                continue
            key = ("SEC", section, twp, rng, mer)
            out_loc = f"SEC-{section}-{twp}-{rng}-{mer}"
        else:
            continue

        def g(c):
            return row[c] if c is not None and c < len(row) else None

        out.append({
            "property_id": property_id,
            "loc": out_loc,
            "loc_raw": lld_s,
            "qtr": qtr, "sec": section, "twp": twp, "rng": rng, "meridian": mer,
            "titled_ac": g(c_size),
            "rm": g(c_muni),
            "cot": g(c_cot),
            "short_legal": g(c_short),
            "crop_2025": g(c_2025),
            "crop_2024": g(c_2024),
            "crop_2023": g(c_2023),
            "_key": key,
            "_polygon": ring,
        })
    wb.close()
    return _aggregate_by_key(out)


def parse_montana(xlsx_path: Path) -> list[dict]:
    """
    Montana.xlsx combines 3 properties. Property assignment uses the
    'PropertyName' column if present, otherwise the county + township
    clustering. Columns observed:
      OBJECTID | PARCELID | COUNTYCD | CountyName | CountyAbbr | GISAcres |
      TaxYear | PropertyID | AssessmentCode | Township ('04 S') |
      Range ('31 E') | Section (26)
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    # First sheet has all three properties.
    ws = wb[wb.sheetnames[0]]
    hdr_row, hdr = _find_header_row(ws, {"township", "range", "section"})
    if hdr_row < 0:
        wb.close()
        return []

    def col(*keys):
        for k in keys:
            if k in hdr:
                return hdr[k]
        return None

    c_twp = col("township")
    c_rng = col("range")
    c_sec = col("section")
    c_acres = col("gisacres", "acres")
    c_county = col("countyname", "county")
    c_county_cd = col("countycd")
    c_parcel_id = col("parcelid")
    c_assessment = col("assessmentcode")
    c_property_name = col("propertyname", "property name", "property")

    out = []
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        tw = row[c_twp] if c_twp is not None and c_twp < len(row) else None
        rn = row[c_rng] if c_rng is not None and c_rng < len(row) else None
        sc = row[c_sec] if c_sec is not None and c_sec < len(row) else None
        if not (tw and rn and sc):
            continue
        # Parse "04 S" → (4, S)
        try:
            tw_s = str(tw).strip().split()
            rn_s = str(rn).strip().split()
            tw_n = int(tw_s[0])
            tw_d = tw_s[1] if len(tw_s) > 1 else "N"
            rn_n = int(rn_s[0])
            rn_d = rn_s[1] if len(rn_s) > 1 else "E"
            section = int(sc)
        except (ValueError, IndexError):
            continue

        try:
            ring = mt_section_polygon(section, tw_n, tw_d, rn_n, rn_d)
        except ValueError:
            continue

        def g(c):
            return row[c] if c is not None and c < len(row) else None

        # data.js groups all 3 MT properties (Fly Creek, Camp 1, Camp 4)
        # under the single id "montana". Match that grouping so the fill-color
        # match-expression and property drawer work consistently. We retain
        # the county in 'county' for UI use.
        property_id = "montana"

        loc_str = f"T{tw_n:02d}{tw_d}-R{rn_n:02d}{rn_d}-S{section:02d}"
        out.append({
            "property_id": property_id,
            "loc": loc_str,
            "loc_raw": f"Sec {section}, T{tw_n} {tw_d}, R{rn_n} {rn_d}",
            "qtr": None, "sec": section, "twp": tw_n, "rng": rn_n,
            "meridian": "MT-PM", "twp_dir": tw_d, "rng_dir": rn_d,
            "titled_ac": g(c_acres),
            "county": g(c_county),
            "county_cd": g(c_county_cd),
            "parcel_id": g(c_parcel_id),
            "assessment": g(c_assessment),
            "_key": (tw_n, tw_d, rn_n, rn_d, section),
            "_polygon": ring,
        })
    wb.close()
    return _aggregate_by_key(out)


# ── Driver ─────────────────────────────────────────────────────────────────
PROPERTY_FILE_MAP = {
    "admiral":       ("Admiral.xlsx",       parse_sk_standard),
    "calderbank":    ("Calderbank.xlsx",    parse_sk_standard),
    "hafford":       ("Hafford.xlsx",       parse_sk_standard),
    "kamsack":       ("Kamsack.xlsx",       parse_sk_standard),
    "outlook":       ("Outlook.xlsx",       parse_sk_standard),
    "ponteix":       ("Ponteix.xlsx",       parse_sk_standard),
    "prince-albert": ("Prince_Albert.xlsx", parse_sk_standard),
    "raymore":       ("Raymore.xlsx",       parse_sk_standard),
    "vanguard":      ("Vanguard.xlsx",      parse_sk_standard),
    "wymark":        ("Wymark.xlsx",        parse_sk_standard),
    "eddystone":     ("Eddystone.xlsx",     parse_eddystone),
    # The_Pas: Red River parish lots — not resolvable by DLS math.
    # Montana: handled separately, three properties in one XLSX.
}


def build_feature_collection(data_dir: Path) -> dict:
    features: list[dict] = []
    stats: dict[str, int] = {}
    calibration_path = Path(__file__).resolve().parent / "quarter_geometry_calibration.json"
    calibration = load_geometry_calibration(calibration_path)

    print(f"calibration file: {calibration_path}")
    if not calibration_path.exists():
        print("  using identity calibration (file not found)")

    for pid, (fname, parser) in PROPERTY_FILE_MAP.items():
        fpath = data_dir / fname
        if not fpath.exists():
            print(f"  [skip] {pid}: missing {fname}", file=sys.stderr)
            continue
        try:
            records = parser(fpath, pid)
        except Exception as e:
            print(f"  [err]  {pid}: {e}", file=sys.stderr)
            continue
        prop_cal = calibration_for_property(calibration, pid)
        if not is_identity_calibration(prop_cal):
            print(
                f"  [cal]  {pid}: east={prop_cal['east_m']:.1f} m, "
                f"north={prop_cal['north_m']:.1f} m, "
                f"scale_x={prop_cal['scale_x']:.6f}, "
                f"scale_y={prop_cal['scale_y']:.6f}, "
                f"rotate={prop_cal['rotate_deg']:.4f} deg"
            )
            records = apply_property_calibration(records, prop_cal)
        stats[pid] = len(records)
        for rec in records:
            ring = rec.pop("_polygon")
            rec.pop("_key", None)
            # JSON doesn't serialize datetimes / etc.; coerce problematic values.
            props = {k: _safe(v) for k, v in rec.items()}
            features.append({
                "type": "Feature",
                "geometry": {"type": "Polygon", "coordinates": [ring]},
                "properties": props,
            })

    # Montana: one file, three properties
    mt_path = data_dir / "Montana.xlsx"
    if mt_path.exists():
        try:
            mt_records = parse_montana(mt_path)
        except Exception as e:
            print(f"  [err]  montana: {e}", file=sys.stderr)
            mt_records = []
        prop_cal = calibration_for_property(calibration, "montana")
        if not is_identity_calibration(prop_cal):
            print(
                f"  [cal]  montana: east={prop_cal['east_m']:.1f} m, "
                f"north={prop_cal['north_m']:.1f} m, "
                f"scale_x={prop_cal['scale_x']:.6f}, "
                f"scale_y={prop_cal['scale_y']:.6f}, "
                f"rotate={prop_cal['rotate_deg']:.4f} deg"
            )
            mt_records = apply_property_calibration(mt_records, prop_cal)
        # All 3 MT properties aggregated under "montana" id to match data.js.
        for rec in mt_records:
            pid = rec["property_id"]
            stats[pid] = stats.get(pid, 0) + 1
            ring = rec.pop("_polygon")
            rec.pop("_key", None)
            props = {k: _safe(v) for k, v in rec.items()}
            features.append({
                "type": "Feature",
                "geometry": {"type": "Polygon", "coordinates": [ring]},
                "properties": props,
            })

    print("\n  parcels per property:")
    for pid, n in sorted(stats.items(), key=lambda kv: -kv[1]):
        print(f"    {pid:<20} {n:>4}")
    print(f"    {'TOTAL':<20} {sum(stats.values()):>4}")

    return {"type": "FeatureCollection", "features": features}


def _safe(v):
    """Coerce pandas/openpyxl values to JSON-friendly primitives."""
    if v is None:
        return None
    if isinstance(v, (str, int, float, bool)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        return v
    # Dates, Decimals, etc → string.
    return str(v)


def main():
    here = Path(__file__).resolve().parent
    data_dir = Path("G:/My Drive/Agriculture/Monette")
    out_path = here.parent / "quarters.geojson"

    print(f"source: {data_dir}")
    print(f"output: {out_path}\n")

    fc = build_feature_collection(data_dir)

    out_path.write_text(json.dumps(fc, separators=(",", ":")), encoding="utf-8")
    print(f"\nWrote {len(fc['features'])} features -> {out_path}")
    print(f"       {out_path.stat().st_size / 1024:.1f} KB")


if __name__ == "__main__":
    main()

"""
audit_eddystone.py — One-off audit of Eddystone.xlsx to compare against
the data.js row (currently titled:21972, parcels:178). Mirrors the
earlier Hafford audit so third-party landlords are surfaced.

Usage:
    python audit_eddystone.py
"""
from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path

import openpyxl

XLSX = Path(r"G:\My Drive\Agriculture\Monette\Eddystone.xlsx")

# Same MB compact pattern the geojson builder uses, so our "unique parcel"
# counts line up with quarters.geojson.
MB_COMPACT_RE = re.compile(
    r"^\s*([NS][EW])\s*(\d{1,2})\s*-\s*(\d{1,2})\s*-\s*(\d{1,2})\s*W\s*$",
    re.IGNORECASE,
)
MB_SECTION_RE = re.compile(
    r"^\s*(?:[NS]?)?(\d{1,2})\s*-\s*(\d{1,2})\s*-\s*(\d{1,2})\s*W\s*$",
    re.IGNORECASE,
)


def _norm_header(v):
    if v is None:
        return ""
    return str(v).strip().lower()


def _find_header_row(ws, required_keys: set[str]):
    """Scan first 10 rows for one that contains all required keys."""
    for r in range(1, 11):
        row = [ _norm_header(c) for c in next(ws.iter_rows(min_row=r, max_row=r, values_only=True)) ]
        row_set = set(row)
        if required_keys.issubset(row_set):
            hdr = {v: i for i, v in enumerate(row) if v}
            return r, hdr
    return -1, {}


def normalize_owner(name: str) -> str:
    """Strip trailing punctuation + collapse whitespace so tiny variants collapse."""
    if not name:
        return ""
    s = re.sub(r"\s+", " ", name.strip())
    s = s.rstrip(",.;")
    return s


def classify_owner(owner: str) -> str:
    u = owner.upper()
    if "MONETTE" in u:
        return "monette"
    return "third-party"


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    print(f"Sheet names: {wb.sheetnames}")
    ws = wb[wb.sheetnames[0]]

    # Print header row candidates
    print("\n--- First 5 rows (raw) ---")
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        print(f"row {r_idx}: {list(row)[:15]}")

    hdr_row, hdr = _find_header_row(ws, {"lld"})
    print(f"\nHeader row: {hdr_row}")
    print(f"Columns seen: {list(hdr.keys())}")

    def col(*keys):
        for k in keys:
            if k in hdr:
                return hdr[k]
        return None

    c_lld     = col("lld")
    c_short   = col("short legal description")
    c_size    = col("size", "titled acres", "titled ac")
    # "Owner" (col B) is the true legal-title owner; "Owner Name" (col D)
    # is a sanitized display label that redacts third parties.
    c_owner   = col("owner")
    c_owner_display = col("owner name")
    c_muni    = col("municipality", "rm")
    c_cot     = col("c of t", "title no.", "title no")
    c_civic   = col("civic address")
    c_roll    = col("roll no.", "roll no")
    c_community = col("community")
    c_propname = col("property name")
    c_2025    = col(2025, "2025")
    c_2024    = col(2024, "2024")
    c_2023    = col(2023, "2023")
    c_soil    = col("soil class", "soil", "soil type")
    c_culti   = col("implied cultivated", "cultivated", "cult")
    c_waste   = col("waste acres", "waste")
    c_assess  = col("assessment", "assessed", "assessed value")

    print(f"\nResolved columns:")
    for name, idx in [("lld", c_lld),("short_legal", c_short),("size", c_size),("owner", c_owner),
                      ("muni", c_muni),("cot", c_cot),("2025", c_2025),("2024", c_2024),("2023", c_2023),
                      ("soil", c_soil),("cultivated", c_culti),("waste", c_waste),("assessment", c_assess)]:
        print(f"  {name:15s}: col idx {idx}")

    # ── Walk all data rows ──
    rows = []
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        def g(c):
            return row[c] if c is not None and c < len(row) else None

        lld_raw = g(c_lld)
        if not isinstance(lld_raw, str):
            continue
        lld = lld_raw.strip()
        if not lld:
            continue

        rows.append({
            "lld": lld,
            "owner": normalize_owner(g(c_owner) or ""),
            "owner_display": normalize_owner(g(c_owner_display) or ""),
            "size": g(c_size),
            "muni": g(c_muni),
            "cot": g(c_cot),
            "civic": g(c_civic),
            "roll": g(c_roll),
            "community": g(c_community),
            "propname": g(c_propname),
            "c2025": g(c_2025),
            "c2024": g(c_2024),
            "c2023": g(c_2023),
            "soil": g(c_soil),
            "cultivated": g(c_culti),
            "waste": g(c_waste),
            "assess": g(c_assess),
        })

    wb.close()
    print(f"\nTotal non-blank rows with LLD: {len(rows)}")

    # ── Aggregate raw (one row per C of T / owner / LLD) ──
    print(f"\n=== RAW ROW TOTALS (every row in XLSX) ===")
    raw_size = sum(float(r["size"]) for r in rows if isinstance(r["size"], (int, float)))
    print(f"  Sum of Size (titled acres) across every row: {raw_size:,.2f}")

    # ── Aggregate by unique LLD (this is what matches quarters.geojson) ──
    by_lld = defaultdict(lambda: {"rows":[], "owners":set(), "size":0.0})
    for r in rows:
        m = MB_COMPACT_RE.match(r["lld"]) or MB_SECTION_RE.match(r["lld"])
        key = r["lld"].upper()  # simple key for audit purposes
        b = by_lld[key]
        b["rows"].append(r)
        b["owners"].add(r["owner"])
        if isinstance(r["size"], (int, float)):
            b["size"] += float(r["size"])

    uniq = len(by_lld)
    uniq_size = sum(b["size"] for b in by_lld.values())
    print(f"\n=== UNIQUE LLD TOTALS (aggregated) ===")
    print(f"  Unique LLD parcels: {uniq}")
    print(f"  Sum of Size across unique LLDs: {uniq_size:,.2f}")

    # ── Owner breakdown ──
    print(f"\n=== OWNER BREAKDOWN (raw rows) ===")
    by_owner = defaultdict(lambda: {"count":0, "size":0.0, "llds":set()})
    for r in rows:
        o = r["owner"] or "(blank)"
        by_owner[o]["count"] += 1
        by_owner[o]["llds"].add(r["lld"].upper())
        if isinstance(r["size"], (int, float)):
            by_owner[o]["size"] += float(r["size"])

    for owner, stats in sorted(by_owner.items(), key=lambda kv: -kv[1]["size"]):
        role = classify_owner(owner)
        print(f"  [{role:11s}] {owner or '(blank)':60s}  rows={stats['count']:4d}  uniqLLDs={len(stats['llds']):4d}  size={stats['size']:10,.2f}")

    # ── Monette vs third-party totals ──
    print(f"\n=== MONETTE vs THIRD-PARTY (raw row sums) ===")
    tot_monette = sum(stats["size"] for o, stats in by_owner.items() if classify_owner(o) == "monette")
    tot_third   = sum(stats["size"] for o, stats in by_owner.items() if classify_owner(o) != "monette")
    print(f"  Monette-entity rows titled acres : {tot_monette:10,.2f}")
    print(f"  Third-party rows titled acres    : {tot_third:10,.2f}")
    print(f"  Total (sum)                      : {tot_monette + tot_third:10,.2f}")

    # ── Soil / crop rollups ──
    print(f"\n=== SOIL CLASS ROLLUP (if present) ===")
    soils = defaultdict(float)
    for r in rows:
        s = r["soil"]
        sz = r["size"] if isinstance(r["size"], (int, float)) else 0.0
        if s is None:
            continue
        soils[str(s).strip()] += sz
    if soils:
        for s, ac in sorted(soils.items(), key=lambda kv: -kv[1]):
            print(f"  {s:20s}  {ac:10,.2f}")
    else:
        print("  (no soil column found)")

    def rollup_crops(col_key: str, label: str):
        print(f"\n=== {label} ===")
        buckets = defaultdict(float)
        for r in rows:
            c = r[col_key]
            if c is None:
                continue
            crop = str(c).strip()
            if not crop or crop.lower() in ("none", "n/a", "na", "-"):
                continue
            sz = r["size"] if isinstance(r["size"], (int, float)) else 0.0
            buckets[crop] += sz
        for crop, ac in sorted(buckets.items(), key=lambda kv: -kv[1])[:12]:
            print(f"  {crop:30s}  {ac:10,.2f}")
        return buckets

    rollup_crops("c2025", "CROPS 2025 (acres)")
    rollup_crops("c2024", "CROPS 2024 (acres)")

    # ── Cultivated / waste / assessment roll-ups ──
    print(f"\n=== CULTIVATED / WASTE / ASSESSMENT (raw rows) ===")
    tot_culti = sum(float(r["cultivated"]) for r in rows if isinstance(r["cultivated"], (int, float)))
    tot_waste = sum(float(r["waste"]) for r in rows if isinstance(r["waste"], (int, float)))
    tot_assess = sum(float(r["assess"]) for r in rows if isinstance(r["assess"], (int, float)))
    print(f"  Cultivated sum : {tot_culti:,.2f}")
    print(f"  Waste sum      : {tot_waste:,.2f}")
    print(f"  Assessment sum : ${tot_assess:,.2f}")

    # ── Distinct values in the Owner column and the Owner Name column ──
    print(f"\n=== DISTINCT 'Owner' (legal) VALUES ===")
    own_counts = defaultdict(int)
    for r in rows:
        own_counts[r["owner"] or "(blank)"] += 1
    for v, c in sorted(own_counts.items(), key=lambda kv: -kv[1]):
        print(f"  {v!r:80s}  rows={c}")

    print(f"\n=== DISTINCT 'Owner Name' (display) VALUES ===")
    disp_counts = defaultdict(int)
    for r in rows:
        disp_counts[r["owner_display"] or "(blank)"] += 1
    for v, c in sorted(disp_counts.items(), key=lambda kv: -kv[1]):
        print(f"  {v!r:80s}  rows={c}")

    # ── Third-party LLD sample (so we can eyeball who the land is ──
    print(f"\n=== THIRD-PARTY OWNER LLD SAMPLES (top owners by acreage) ===")
    top_third = sorted(
        ((o, s) for o, s in by_owner.items() if classify_owner(o) != "monette"),
        key=lambda kv: -kv[1]["size"],
    )[:8]
    for owner, stats in top_third:
        sample_llds = sorted(stats["llds"])[:6]
        muni_set = sorted({r["muni"] for r in rows if r["owner"] == owner and r["muni"]})
        print(f"\n  OWNER: {owner}")
        print(f"    rows={stats['count']}  uniqLLDs={len(stats['llds'])}  size={stats['size']:,.2f}")
        print(f"    RMs: {muni_set}")
        print(f"    First LLDs: {sample_llds}")


if __name__ == "__main__":
    main()

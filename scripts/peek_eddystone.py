"""Quick diagnostic: show the contrast between Monette-owned and
3rd-Party-Vendor rows in Eddystone.xlsx — full row dumps for a sample."""
from __future__ import annotations
from pathlib import Path
import openpyxl

wb = openpyxl.load_workbook(r"G:\My Drive\Agriculture\Monette\Eddystone.xlsx",
                            data_only=True, read_only=True)
ws = wb["13-Eddystone"]
rows = list(ws.iter_rows(values_only=True))
header = rows[1]
print("HEADER:")
for i, h in enumerate(header):
    print(f"  col {i:2d}: {h!r}")

data = rows[2:]
print(f"\nTotal data rows: {len(data)}")

monette_idx, third_idx = [], []
for i, r in enumerate(data):
    own_name = r[3]
    if own_name == "Monette Farms":
        monette_idx.append(i)
    elif own_name == "3rd Party Vendor":
        third_idx.append(i)

print(f"\n--- FIRST 5 MONETTE ROWS ---")
for i in monette_idx[:5]:
    r = data[i]
    print(f"Row {i+3}:")
    for col, h in enumerate(header):
        if h is None: continue
        print(f"  {str(h):40s} = {r[col]!r}")
    print()

print(f"\n--- FIRST 5 '3rd Party Vendor' ROWS ---")
for i in third_idx[:5]:
    r = data[i]
    print(f"Row {i+3}:")
    for col, h in enumerate(header):
        if h is None: continue
        print(f"  {str(h):40s} = {r[col]!r}")
    print()

# Also inspect the buildings sheet for owner names
if "13-Eddystone_Bldgs" in wb.sheetnames:
    print(f"\n--- BUILDINGS SHEET HEADER ---")
    ws2 = wb["13-Eddystone_Bldgs"]
    for r_idx, row in enumerate(ws2.iter_rows(min_row=1, max_row=6, values_only=True), start=1):
        print(f"row {r_idx}: {list(row)[:15]}")

# Offer price totals on the Monette sheet
print(f"\n--- OFFER PRICE TOTALS ---")
price_monette = sum(float(r[0]) for i,r in enumerate(data) if data[i][3]=="Monette Farms" and isinstance(r[0],(int,float)))
price_third   = sum(float(r[0]) for i,r in enumerate(data) if data[i][3]=="3rd Party Vendor" and isinstance(r[0],(int,float)))
print(f"  Sum of offer prices on Monette rows: ${price_monette:,.2f}")
print(f"  Sum of offer prices on 3rd-Party rows: ${price_third:,.2f}")
print(f"  Monette rows with offer price set: {sum(1 for i,r in enumerate(data) if data[i][3]=='Monette Farms' and isinstance(r[0],(int,float)))}")
print(f"  3rd-Party rows with offer price set: {sum(1 for i,r in enumerate(data) if data[i][3]=='3rd Party Vendor' and isinstance(r[0],(int,float)))}")

wb.close()

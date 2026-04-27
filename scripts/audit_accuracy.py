"""
Audit current quarters.geojson for accuracy issues:
- Per-property centroid vs. data.js anchor (flags big misalignments)
- Per-property count vs. data.js claim (flags over/undercount)
- Dropped XLSX rows (parcels that didn't parse)
- Duplicates (same loc parsed twice)
- Pull a raw sample of Hafford and Montana to see why counts differ
"""
from __future__ import annotations
import json
import math
import re
from pathlib import Path
from collections import Counter, defaultdict

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
GEOJSON = ROOT / "quarters.geojson"
DATA_DIR = Path("G:/My Drive/Agriculture/Monette")

DATA_JS_ANCHORS = {
    "admiral":       (49.43, -108.03, 16),
    "calderbank":    (50.28, -106.78, 141),
    "hafford":       (52.74, -107.37, 27),
    "kamsack":       (51.56, -101.89, 107),
    "outlook":       (51.50, -107.06, 26),
    "ponteix":       (49.77, -107.43, 130),
    "prince-albert": (53.20, -105.75, 22),
    "raymore":       (51.39, -104.64, 125),
    "vanguard":      (49.92, -107.24, 111),
    "wymark":        (50.11, -107.63, 131),
    "eddystone":     (51.13,  -99.47, 178),
    "the-pas":       (53.82, -101.25, 147),
    "montana":       (48.10, -106.50,  36),
}

fc = json.loads(GEOJSON.read_text(encoding="utf-8"))

by_pid = defaultdict(list)
for feat in fc["features"]:
    by_pid[feat["properties"]["property_id"]].append(feat)

print("=" * 78)
print(f"{'property':<15} {'gj_cnt':>6}  {'dj_cnt':>6}  {'ctr_lat':>9}  {'ctr_lng':>9}  {'anchor_lat':>10}  {'anchor_lng':>10}  {'dlat':>6}  {'dlng':>6}")
print("-" * 78)
for pid, (a_lat, a_lng, a_cnt) in DATA_JS_ANCHORS.items():
    feats = by_pid.get(pid, [])
    n = len(feats)
    if n == 0:
        print(f"{pid:<15} {n:>6}  {a_cnt:>6}   ---        ---        {a_lat:>10.3f}  {a_lng:>10.3f}  ---    ---")
        continue
    lats, lngs = [], []
    for f in feats:
        ring = f["geometry"]["coordinates"][0]
        cx = sum(p[0] for p in ring[:-1]) / 4
        cy = sum(p[1] for p in ring[:-1]) / 4
        lngs.append(cx)
        lats.append(cy)
    ctr_lat = sum(lats) / len(lats)
    ctr_lng = sum(lngs) / len(lngs)
    dlat = ctr_lat - a_lat
    dlng = ctr_lng - a_lng
    flag = "  !!" if (abs(dlat) > 0.5 or abs(dlng) > 1.0) else ""
    cnt_flag = "  !!" if abs(n - a_cnt) > max(3, a_cnt * 0.1) else ""
    print(f"{pid:<15} {n:>6}  {a_cnt:>6}  {ctr_lat:>9.3f}  {ctr_lng:>9.3f}  {a_lat:>10.3f}  {a_lng:>10.3f}  {dlat:>+6.2f}  {dlng:>+6.2f}{flag}{cnt_flag}")

print()
print("=" * 78)
print("Duplicate locs per property:")
print("-" * 78)
for pid, feats in by_pid.items():
    locs = [f["properties"].get("loc") for f in feats]
    dupes = [(loc, cnt) for loc, cnt in Counter(locs).items() if cnt > 1 and loc]
    if dupes:
        print(f"  {pid}: {len(dupes)} distinct locs duplicated; top 5:")
        for loc, cnt in sorted(dupes, key=lambda x: -x[1])[:5]:
            print(f"    {loc:>24}  x{cnt}")

# Raw XLSX inspection for Hafford (expected=27 per data.js but 161 in geojson)
print()
print("=" * 78)
print("Hafford raw XLSX sample (header + first 8 rows where Land Locations is set):")
print("-" * 78)
wb = openpyxl.load_workbook(DATA_DIR / "Hafford.xlsx", data_only=True, read_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
# find header row
hdr_idx = None
for i, r in enumerate(rows):
    if r and any(isinstance(c, str) and c.strip().lower() == "land locations" for c in r):
        hdr_idx = i
        break
if hdr_idx is not None:
    hdr = [str(c).strip() if c else "" for c in rows[hdr_idx]]
    loc_col = next(i for i, h in enumerate(hdr) if h.lower() == "land locations")
    title_col = next((i for i, h in enumerate(hdr) if h.lower() == "title"), None)
    parcel_col = next((i for i, h in enumerate(hdr) if "parcel" in h.lower()), None)
    print(f"  headers at row {hdr_idx+1}: loc@{loc_col}, title@{title_col}, parcel@{parcel_col}")
    print(f"  total rows after header: {len(rows) - hdr_idx - 1}")
    distinct_locs = set()
    distinct_titles = set()
    for r in rows[hdr_idx+1:]:
        if loc_col < len(r) and isinstance(r[loc_col], str):
            distinct_locs.add(r[loc_col].strip())
        if title_col is not None and title_col < len(r) and r[title_col]:
            distinct_titles.add(str(r[title_col]).strip())
    print(f"  distinct Land Locations: {len(distinct_locs)}")
    print(f"  distinct Titles:         {len(distinct_titles)}")
wb.close()

# Montana — expected 36 in data.js, 220 in geojson
print()
print("=" * 78)
print("Montana raw XLSX sample:")
print("-" * 78)
wb = openpyxl.load_workbook(DATA_DIR / "Montana.xlsx", data_only=True, read_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
hdr = [str(c).strip() if c else "" for c in rows[0]]
print(f"  headers: {hdr}")
print(f"  total rows after header: {len(rows) - 1}")
# distinct (twp, rng, sec)
try:
    c_twp = next(i for i, h in enumerate(hdr) if h.lower() == "township")
    c_rng = next(i for i, h in enumerate(hdr) if h.lower() == "range")
    c_sec = next(i for i, h in enumerate(hdr) if h.lower() == "section")
    c_tax = next((i for i, h in enumerate(hdr) if "taxyear" in h.lower() or h.lower() == "taxyear"), None)
    distinct = set()
    by_year = Counter()
    for r in rows[1:]:
        if c_twp < len(r) and c_rng < len(r) and c_sec < len(r):
            distinct.add((r[c_twp], r[c_rng], r[c_sec]))
            if c_tax is not None and c_tax < len(r):
                by_year[r[c_tax]] += 1
    print(f"  distinct (Township, Range, Section) combos: {len(distinct)}")
    print(f"  rows by TaxYear: {dict(by_year)}")
except StopIteration:
    print("  column lookup failed")
wb.close()

# Eddystone — expected 178 in data.js, 174 in geojson; what didn't parse?
print()
print("=" * 78)
print("Eddystone unparseable rows:")
print("-" * 78)
MB_RE = re.compile(r"^\s*(NE|NW|SE|SW)(\d{1,2})-(\d{1,3})-(\d{1,3})(W[1-6]?)?\s*$", re.I)
wb = openpyxl.load_workbook(DATA_DIR / "Eddystone.xlsx", data_only=True, read_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
hdr_idx = None
for i, r in enumerate(rows):
    if r and any(isinstance(c, str) and c.strip().lower() == "lld" for c in r):
        hdr_idx = i
        break
if hdr_idx is not None:
    hdr = [str(c).strip() if c else "" for c in rows[hdr_idx]]
    c_lld = next(i for i, h in enumerate(hdr) if h.lower() == "lld")
    total = 0
    bad = []
    for r in rows[hdr_idx+1:]:
        if c_lld < len(r) and isinstance(r[c_lld], str) and r[c_lld].strip():
            total += 1
            if not MB_RE.match(r[c_lld].strip()):
                bad.append(r[c_lld].strip())
    print(f"  total non-empty LLD rows: {total}")
    print(f"  unparseable: {len(bad)}")
    for b in bad[:12]:
        print(f"    {b!r}")
wb.close()
